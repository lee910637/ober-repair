"""
建立 17108A-5-120 維修診斷原型的來源 Excel（Excel 為主、App 唯讀）。
內容依據 0_concept/TC_17108A-5-120_MaintenanceManual.docx 與
0_concept/17108-5-120 A板簡易查修說明.pptx 轉錄整理。

執行方式：../.venv/bin/python make_source_xlsx.py
輸出：17108A-5-120_診斷資料.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

HEADER_FILL = PatternFill(start_color="1F4E5F", end_color="1F4E5F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def write_sheet(ws, headers, rows):
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append(r)
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        maxlen = max([len(str(headers[col - 1]))] + [len(str(r[col - 1])) for r in rows] + [10])
        ws.column_dimensions[letter].width = min(maxlen + 2, 60)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# 1. 異常資訊 (Faults) —— 對應手冊第 3 章「全機保護訊息」表 + 第 5 章查修方向
# ---------------------------------------------------------------------------
ws1 = wb.active
ws1.title = "異常資訊"
faults_headers = ["Fault_ID", "名稱", "關鍵字", "類別", "優先等級", "是否已建置", "起始節點", "手冊章節", "說明"]
faults_rows = [
    ["F_CALI", "Calibration Error 校正值異常", "校正值異常,Slope,Calibration Out of Range", "校正", "高", "否", "", "4.8節", "校正值 Slope>20%，判定為 Calibration Out of Range。"],
    ["F_OCP", "OCP／OVP 過電流過電壓保護", "OCP,OVP,過電流,過電壓,Input Voltage Protection", "保護", "高", "是", "OCP_1", "4.1節", "Iout>134A 或 Vout>5.8V 或 Vin>14V 時觸發。"],
    ["F_OTP", "OTP 單機環溫過溫保護", "OTP,過溫,溫度異常", "溫度", "高", "否", "", "4.2節", "單機環溫 >100°C 觸發過溫保護。"],
    ["F_CONTACT", "接觸錯誤／感應異常保護", "接觸錯誤,感應異常,Vmi,Vmo", "量測", "高", "否", "", "4.3節", "ABS(Vmi-Vmo)/I 超出設定值，Sense/Drive 線路或量測電路異常。"],
    ["F_FAN", "FAN 風扇異常保護", "Fan,風扇,異常,鎖轉,轉速", "散熱", "中", "是", "FAN_1", "4.5節", "風扇電源異常或風扇本體損壞造成 Fan Lock 保護。"],
    ["F_NET", "網路通訊異常", "網路,LAN,IP,通訊異常", "通訊", "中", "否", "", "4.4節", "PC 與單機之間 LAN 通訊異常，無法連線。"],
    ["F_SYSCFG", "System Config Error", "System Config,模組錯誤,12V,13V", "控制", "中", "否", "", "4.6節", "前級轉換器(AC-DC)異常或通道數/Slot數設定錯誤。"],
    ["F_ADC", "ADC Error 讀值異常", "ADC,讀值異常,固定0V,固定0A", "量測", "高", "否", "", "4.7節", "電壓/電流讀值固定不跳動，疑似 ADC 或 A 板異常。"],
    ["F_NOOUTPUT", "無輸出電壓，DVM無讀值", "無輸出電壓,DVM無讀值,CV4V 0V 0V", "校驗證查修", "高", "是", "NOV_1", "5.2.1節", "執行 8000 程式校驗證時，設定電壓後完全無輸出，內部讀值與 DVM 實測值皆為 0V。"],
]
write_sheet(ws1, faults_headers, faults_rows)

# ---------------------------------------------------------------------------
# 2. 確認資訊 (Checks) —— 可重複使用的原子檢測步驟
# ---------------------------------------------------------------------------
ws2 = wb.create_sheet("確認資訊")
checks_headers = ["Check_ID", "檢查部位", "檢查方法", "正常條件/標準值", "單位", "分類"]
checks_rows = [
    # --- OCP ---
    ["C_OCP_PERIPH", "OCP/OVP/Input_OVP/比較器/Ref voltage 週邊元件",
     "使用三用電錶量測 R726/R868、R724/R866、C710/C850、R744/R888、R745/R887、C708/C848 等週邊電阻電容",
     "R726/R868=75KΩ±1%; R724/R866=100KΩ±1%; C710/C850=0.1uF±10%; R744/R888=10KΩ±1%; R745/R887=10KΩ±1%; C708/C848=1uF±10%", "Ω/uF", "保護"],
    ["C_OCP_U602", "U602 Pin5~Pin7", "以 DMM 量測 U602 Pin5、Pin6、Pin7 電壓", "正常: Pin6=3.33V,Pin5=2.5V,Pin7=0V；OCP觸發: Pin6≈0.7V,Pin5=2.5V,Pin7=3.33V", "V", "保護"],
    ["C_OCP_R603R605", "R603 / R605", "量測 R603 右邊電壓、R605 右邊電壓", "R603右邊應為10V；R605右邊應為5V", "V", "保護"],
    ["C_OCP_DGROUP", "D602 / D607 / D608 (input_OCP_bar / Fault_02)", "以 DMM 量測二極體兩端電壓", "正常應為 HIGH；電壓 <1V 視為 LOW(異常)", "V", "保護"],
    ["C_OCP_INPUTOVP", "U602 Pin1 對 Pin4(GND) / R620", "量測 U602 Pin1 對 GND 電壓；量測 R620 阻抗", "Pin1為HIGH代表Input OVP；R620應為15KΩ±0.1%", "V/Ω", "保護"],
    ["C_OCP_R816", "R816 下方 / R791 下方", "量測 R816(或R791) 下方電壓", "正常應為 HIGH，LOW 代表控制IC(8450)跳OCP/OV", "V", "保護"],
    ["C_OCP_C849C850", "C849(OVP) / C850(OCP) / C709(OVP) / C710(OCP)", "量測電容電壓", "電壓 >1.25V 代表該相觸發對應保護", "V", "保護"],
    ["C_OCP_VDROP", "R890/R891 下方壓差 / R747/R748 下方壓差", "量測電阻下方壓差", "應等於電池電壓", "V", "保護"],
    # --- FAN ---
    ["C_FAN_DEBRIS", "風扇葉片/入風口", "目視+手動檢查是否有異物、積塵造成堵轉", "無異物、無堵轉", "-", "散熱"],
    ["C_FAN_BODY", "風扇本體", "通電測試風扇是否正常運轉(有無異音、轉速是否明顯不足)", "運轉順暢、無異音", "-", "散熱"],
    ["C_FAN_POWER", "風扇線材(FT板) / 風扇電源(S板→FT板)", "量測風扇電源迴路電壓", "供電正常，無斷路或電壓異常", "V", "散熱"],
    # --- 無輸出電壓 ---
    ["C_NOV_MOSRELAY_V", "MOS Relay 前端", "量測 MOS Relay 前端是否有電壓", "依機種規格應有電壓", "V", "功率"],
    ["C_NOV_GS_V", "MOS Relay G-S 腳", "量測 MOS Relay Gate-Source 電壓", "依驅動規格，正常應能驅動導通", "V", "功率"],
    ["C_NOV_U101U102", "U101 & U102 PIN1 / PIN3", "量測驅動 IC 輸入輸出電壓", "依機種規格", "V", "功率"],
    ["C_NOV_GS_PWM", "MOS G-S 腳", "以示波器量測 MOS Gate-Source 是否有 PWM 波形", "應有連續 PWM 波形", "-", "功率"],
    ["C_NOV_U708803_1510", "U708/U803 PIN15 & PIN10", "示波器量測 PWM 波形", "應有 PWM 波形", "-", "功率"],
    ["C_NOV_U708803_1", "U708/U803 PIN1", "示波器量測 PWM 波形", "應有 PWM 波形", "-", "功率"],
    ["C_NOV_U707802_6", "U707/U802 PIN6", "示波器量測波形", "應有波形", "-", "功率"],
    ["C_NOV_U707802_5", "U707/U802 PIN5", "示波器量測波形", "應有波形", "-", "功率"],
    ["C_NOV_U706801_2", "U706/U801 PIN2", "示波器量測波形", "應有波形", "-", "功率"],
    ["C_NOV_U706801_5", "U706/U801 PIN5", "DMM量測電壓", "應為 5V", "V", "功率"],
    ["C_NOV_PIN9GROUP", "PIN9 / D703&D803負端 / U705&U804 G腳", "DMM量測各點電壓", "U705/U804 G腳應為 5V", "V", "功率"],
    ["C_NOV_U702805_PINS", "U702/U805 PIN39/40/44/49/60、C766", "DMM量測各腳位電壓", "依機種規格，各點電壓應正常", "V", "功率"],
    ["C_NOV_U1001", "U1001 PIN1 / R1007", "DMM量測電壓", "依機種規格應正常", "V", "數位"],
    ["C_NOV_U322GROUP", "U306 / U307 / U312", "DMM量測各IC電壓", "依機種規格應正常", "V", "數位"],
]
write_sheet(ws2, checks_headers, checks_rows)

# ---------------------------------------------------------------------------
# 3. 判斷節點 (Decision Nodes) —— 決策樹
#    選項下一步：Node_ID，或 "END::結論::建議處置"
# ---------------------------------------------------------------------------
ws3 = wb.create_sheet("判斷節點")
nodes_headers = [
    "Node_ID", "Fault_ID", "Check_ID", "提示文字",
    "選項1文字", "選項1下一步", "選項2文字", "選項2下一步", "選項3文字", "選項3下一步",
]
nodes_rows = [
    # ============================= OCP =============================
    ["OCP_1", "F_OCP", "C_OCP_PERIPH",
     "如尚未啟動電源即出現 OCP，先量測 OCP/OVP/Input_OVP/比較器/Ref voltage 週邊元件，確認是否都在合格範圍內。",
     "有元件數值超出合格範圍", "END::更換超出公差的週邊元件（對照規格表）::依規格表更換異常的電阻/電容",
     "全部元件數值正常", "OCP_2", "", ""],
    ["OCP_2", "F_OCP", "C_OCP_U602",
     "量測 U602 Pin5~Pin7 電壓。正常：Pin6=3.33V、Pin5=2.5V、Pin7=0V；OCP觸發時：Pin6≈0.7V、Pin5=2.5V、Pin7=3.33V。量測結果符合上述哪一種？",
     "符合正常或OCP觸發任一種模式", "OCP_4", "都不符合(讀值異常/中間值)", "OCP_3",
     "", ""],
    ["OCP_3", "F_OCP", "C_OCP_R603R605",
     "量 R603 右邊是否為10V；R605 右邊是否為5V。",
     "皆正常(10V/5V)", "END::邏輯IC(LM393)異常，造成Pin1輸出異常::更換 LM393",
     "任一不正常", "END::PCB layout斷路::依手冊跳線：10V點跳線到C611上方、5V點跳線到R619(下)、Pin5跳線到C540",
     "", ""],
    ["OCP_4", "F_OCP", "C_OCP_DGROUP",
     "量測 D602、D607、D608(input_OCP_bar / Fault_02)。正常應為HIGH，<1V視為LOW(異常)。D608對應Input OVP路徑，D602/D607對應OCP/OVP路徑。",
     "D608 為 LOW", "OCP_5", "D602 或 D607 為 LOW", "OCP_6",
     "皆為 HIGH(正常)", "END::保護迴路偵測本身正常，異常另有原因::建議會同資深工程師覆核控制IC與命令路徑，必要時更換數位板"],
    ["OCP_5", "F_OCP", "C_OCP_INPUTOVP",
     "D608為LOW，代表可能跳 Input OVP。請先回想：OCP是否只在『放電』時才觸發？並量測 U602 Pin1 對 Pin4(GND) 是否為HIGH，以及 R620 阻抗是否為18KΩ。",
     "放電時才觸發 且 Pin1為HIGH", "END::確認為 Input OVP::檢查R620是否18KΩ±0.1%，阻抗異常則更換R620；阻抗正常則為Input OVP偵測電路其他元件異常，需逐一排查",
     "非上述情況", "END::非典型Input OVP案例::建議會同資深工程師覆核電路，勿逕自更換零件",
     "", ""],
    ["OCP_6", "F_OCP", "C_OCP_R816",
     "D602或D607為LOW，量測 R816下(或R791下) 是否同為LOW。",
     "R816(或R791)下方為LOW", "OCP_7", "為HIGH(不同)",
     "END::控制IC(8450)未跳保護::建議另尋訊號路徑異常點，可能為量測電路本身異常", "", ""],
    ["OCP_7", "F_OCP", "C_OCP_C849C850",
     "R816(或R791)為LOW，代表控制IC 8450 跳 OCP/OV。量測 C849(OVP)/C850(OCP)（或另一相 C709(OVP)/C710(OCP)）電壓是否超過1.25V。",
     "C849(或C709) 超過1.25V (OVP)", "OCP_8",
     "C850(或C710) 超過1.25V (OCP)", "END::該相觸發 OCP 過電流保護::檢查功率級電流迴路與MOSFET/電流取樣電路，比對更換相關功率元件",
     "", ""],
    ["OCP_8", "F_OCP", "C_OCP_VDROP",
     "確認為OVP，需量測實際觸發電壓：C849超過1.25V則量R890與R891下方壓差；C709超過1.25V則量R747與R748下方壓差，此壓差應等於電池電壓。",
     "壓差 = 電池電壓(符合預期)", "END::該相觸發 OVP 過電壓保護，屬正常保護動作::請確認電池/測試條件是否超出設備輸出電壓上限(5.8V)，非硬體故障",
     "壓差明顯異常(不等於電池電壓)", "END::電壓量測迴路(Vmo)異常::更換 U702/U805 或相關量測迴路元件",
     "", ""],
    # ============================= FAN =============================
    ["FAN_1", "F_FAN", "C_FAN_DEBRIS",
     "檢查風扇是否有異物或積塵造成堵轉。",
     "有異物/堵轉", "END::風扇堵轉::清除異物/積塵後重新測試",
     "無異物，風扇順暢", "FAN_2", "", ""],
    ["FAN_2", "F_FAN", "C_FAN_BODY",
     "確認無堵轉後，通電檢查風扇本體是否損壞（完全不轉/有異音/轉速明顯不足）。",
     "風扇本體損壞", "END::風扇本體損壞::更換風扇",
     "風扇本體正常", "FAN_3", "", ""],
    ["FAN_3", "F_FAN", "C_FAN_POWER",
     "檢查風扇電源迴路（S板→FT板→風扇）是否正常供電。",
     "電源異常", "END::S板電源異常導致風扇無法正常運轉::更換S板後重新開機確認異常是否解除；若未解除，接續排查FT板",
     "電源正常", "END::風扇本體與線路皆正常，但面板仍顯示異常::判斷為FT板電路異常，更換FT板後開機確認",
     "", ""],
    # ======================= 無輸出電壓 DVM無讀值 =======================
    ["NOV_1", "F_NOOUTPUT", "C_NOV_MOSRELAY_V",
     "不良現象：無輸出電壓，DVM無讀值。查修方向：確認 MOS Relay 前是否有電壓。",
     "有電壓(猜測MOS Relay未導通)", "NOV_1B", "無電壓", "NOV_2", "", ""],
    ["NOV_1B", "F_NOOUTPUT", "C_NOV_GS_V",
     "確認 MOS Relay_GS 腳電壓是否正常。",
     "正常", "END::MOS Relay 零件損壞::更換 MOS Relay",
     "異常", "END::驅動電路(U101/U102)異常::確認 U101/U102 PIN1、PIN3 是否有電壓，異常則更換相關驅動零件",
     "", ""],
    ["NOV_2", "F_NOOUTPUT", "C_NOV_GS_PWM",
     "MOS Relay前無電壓。確認 MOS_GS 腳是否有 PWM 波形。",
     "有 PWM 波形", "END::MOS 可能損壞::更換 MOS",
     "無 PWM 波形", "NOV_3A", "", ""],
    ["NOV_3A", "F_NOOUTPUT", "C_NOV_U708803_1510",
     "確認 U708/U803 PIN15 與 PIN10 是否有 PWM 波形。",
     "有波形", "END::確認MOS背面(功率級)零件是否異常::進一步量測MOS周邊被動元件，異常則更換",
     "無波形", "NOV_3B", "", ""],
    ["NOV_3B", "F_NOOUTPUT", "C_NOV_U708803_1",
     "確認 U708/U803 PIN1 是否有 PWM 波形。",
     "有波形", "END::U708/U803本身可能異常::比對各腳位電壓及波形，異常則更換 U708/U803",
     "無波形", "NOV_4A", "", ""],
    ["NOV_4A", "F_NOOUTPUT", "C_NOV_U707802_6",
     "往 U707/U802 查修：確認 PIN6 是否有波形。",
     "有波形", "END::R757、R824 電阻可能異常::量測後異常則更換",
     "無波形", "NOV_4B", "", ""],
    ["NOV_4B", "F_NOOUTPUT", "C_NOV_U707802_5",
     "確認 U707/U802 PIN5 是否有波形。",
     "有波形", "END::附近零件或U707/U802本身可能異常::比對各腳位電壓與波形，異常則更換相關零件",
     "無波形", "NOV_5A", "", ""],
    ["NOV_5A", "F_NOOUTPUT", "C_NOV_U706801_2",
     "往 U706/U801 查修：確認 PIN2 是否有波形。",
     "有波形", "END::量測 U706_PIN2 與 U707_PIN5 是否導通::不導通表示PIN2到PIN5路徑斷路，需檢修該連接",
     "無波形", "NOV_5B", "", ""],
    ["NOV_5B", "F_NOOUTPUT", "C_NOV_U706801_5",
     "確認 U706/U801 PIN5 是否有5V。",
     "無5V", "END::供電異常::往 U315/U314 查修，確認輸出是否正常，異常則更換",
     "有5V(正常)", "NOV_5C", "", ""],
    ["NOV_5C", "F_NOOUTPUT", "C_NOV_PIN9GROUP",
     "比對 PIN9 電壓(決定輸出電壓大小)、確認 D703/D803 負端電壓、確認 U705/U804 G腳電壓是否為5V。",
     "G腳電壓非5V", "END::供電或訊號路徑異常::往 U316 查修，並同時檢查 U702/U805，異常則更換",
     "G腳電壓為5V(正常)", "NOV_6", "", ""],
    ["NOV_6", "F_NOOUTPUT", "C_NOV_U702805_PINS",
     "往 U702/U805 查修：確認 PIN39、PIN40、PIN44、PIN49、PIN60 電壓，以及 C766 電壓是否正常。",
     "有任一電壓異常", "NOV_7",
     "電壓皆正常", "END::U702/U805本身工作正常但仍無輸出::進一步比對驅動時序，或將 U702/U805 整體模組送修",
     "", ""],
    ["NOV_7", "F_NOOUTPUT", "C_NOV_U1001",
     "往 U1001 查修：確認 PIN1 電壓與 R1007 電壓是否正常。",
     "異常", "END::U1001周邊電路異常::往 U301 查修，確認各腳位電壓、波形，異常則更換U301",
     "皆正常", "NOV_8", "", ""],
    ["NOV_8", "F_NOOUTPUT", "C_NOV_U322GROUP",
     "往 U322 查修：確認 U306、U307、U312 電壓是否正常。",
     "有異常", "END::U306/U307/U312其中之一異常::更換異常的該顆IC",
     "皆無異常", "END::往 U319/U320/U321 查修::進一步比對，異常則更換 U319、U320 或 U321",
     "", ""],
]
write_sheet(ws3, nodes_headers, nodes_rows)

out_path = "17108A-5-120_診斷資料.xlsx"
wb.save(out_path)
print(f"已建立 {out_path}")
print(f"異常資訊: {len(faults_rows)} 筆, 確認資訊: {len(checks_rows)} 筆, 判斷節點: {len(nodes_rows)} 筆")
